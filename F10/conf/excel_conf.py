#!/usr/bin/env python

# coding: utf-8 

# @Time : 2022/6/6 21:38
'''
Excel 单元格写入时的样式
'''
# @Author : xumenghan
from openpyxl.styles import PatternFill, Font


def write_result(cell, row, column, status=1):
    '''

    :param cell:
    :param row:
    :param column:
    :param status: 1 为true  2 为false
    :return:
    '''
    if status == 1:
        cell(row=row, column=column).value = 'Pass'
        # 单元格样式定义：绿色+加粗
        cell(row=row, column=column).fill = PatternFill('solid', fgColor='AACF91')
        cell(row=row, column=column).font = Font(bold=True)
    else:
        cell(row=row, column=column).value = 'Failed'
        # 单元格样式定义：红色+加粗
        cell(row=row, column=column).fill = PatternFill('solid', fgColor='FF0000')
        cell(row=row, column=column).font = Font(bold=True)

    cell(row=row, column=column).font = Font(bold=True)
