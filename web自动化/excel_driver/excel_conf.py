#!/usr/bin/env python

# coding: utf-8 

# @Time : 2022/5/22 17:34

# @Author : Lelsey

from openpyxl.styles import PatternFill, Font


def pass_(cell, row, column):
    # 写入内容
    cell(row=row, column=column).value = 'Pass'
    # 单元格样式定义：绿色加粗
    cell(row=row, column=column).fill = PatternFill('solid', fgColor='AACF91')
    cell(row=row, column=column).font = Font(bold=True)


def fail_(cell, row, column):
    cell(row=row, column=column).value = 'False'
    cell(row=row, column=column).fill = PatternFill('solid', fgColor='FF0000')
    cell(row=row, column=column).font = Font(bold=True)
